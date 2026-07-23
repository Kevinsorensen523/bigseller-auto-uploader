"""Import/export varian lewat Excel (.xlsx) atau CSV, supaya produk dengan
banyak varian (puluhan baris) tidak perlu diketik satu-satu di web form."""

import csv
import io

import openpyxl

COLUMNS = ["value", "sku", "price", "stock"]
COLUMN_ALIASES = {
    "value": {"value", "nilai", "varian", "tipe", "opsi"},
    "sku": {"sku"},
    "price": {"price", "harga"},
    "stock": {"stock", "stok"},
}


def _match_column(header_cell: str) -> str | None:
    key = (header_cell or "").strip().lower()
    for column, aliases in COLUMN_ALIASES.items():
        if key in aliases:
            return column
    return None


def _rows_to_variants(rows: list[list[str]]) -> list[dict]:
    if not rows:
        return []

    header = [_match_column(cell) for cell in rows[0]]
    variants = []
    for row in rows[1:]:
        if not any(str(c).strip() for c in row if c is not None):
            continue
        record = {}
        for column, cell in zip(header, row):
            if column:
                record[column] = str(cell).strip() if cell is not None else ""
        if not record.get("value"):
            continue
        variants.append(
            {
                "value": record.get("value", ""),
                "sku": record.get("sku", ""),
                "price": int(float(record.get("price") or 0)),
                "stock": int(float(record.get("stock") or 0)),
            }
        )
    return variants


def parse_variant_file(file_storage) -> list[dict]:
    """file_storage: objek file dari request.files (Werkzeug FileStorage)."""
    filename = (file_storage.filename or "").lower()
    content = file_storage.read()

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]

    return _rows_to_variants(rows)


def build_template_xlsx() -> io.BytesIO:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Varian"
    sheet.append(COLUMNS)
    sheet.append(["S21", "TG-KA-A08-SAMSUNG-S21", 180000, 1000])
    sheet.append(["S21+", "TG-KA-A08-SAMSUNG-S21PLUS", 180000, 1000])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
